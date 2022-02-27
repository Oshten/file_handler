import re
import openpyxl as xl


def xlsx_handler(file):
    """Обработка файла .xlsx"""
    for workbook in xl.load_workbook(file):
        for col in workbook.iter_cols():
            if col[0].value == 'after':
                # Ищем список L1 или L2
                """из описания задания не понятно что из себя представляет список L1 или L2
                в коде буду исходить из того, что это строчка типа
                '1, 2, 3, 4, 5, 6, 7, 8, 9'.
                Если имеется в виду что-то другое, то в этом месте следует изменить код"""
                for i in range(0, workbook.max_row):
                    if type(col[i].value) == str and re.findall(r'\d+[,\s$]{1,}', col[i].value):
                        list_after = col[i].value
            elif col[0].value == 'before':
                # аналогично ищем для колонки 'before'
                for i in range(0, workbook.max_row):
                    if type(col[i].value) == str and re.findall(r'\d+[,\s$]{1,}', col[i].value):
                        list_before = col[i].value
    return list_after, list_before

def list_handler(list_values):
    """Перобразование строки в список чисел"""
    list_resalt = []
    for value in list_values.split(','):
        list_resalt.append(int(value.replace(',', '').strip()))
    return list_resalt

def list_comparison(list_1, list_2):
    """Сравнение списков
    (количество елементов в списке list_1 должно быть больше количесства элементов в списке list_2)"""
    for value in list_2:
        list_1.remove(value)
    return list_1
